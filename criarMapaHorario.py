from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from tkinter import filedialog, simpledialog, Tk
import win32com.client as win32

sunday_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
saturday_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
holiday_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
weekday_font = Font(color="000000")
sunday_font = Font(color="FFFFFF")
saturday_font = Font(color="000000")
holiday_font = Font(color="FFFFFF")
bold_font = Font(bold=True)
working_values = ["M", "m", "T", "t", "N", "TG", "R"]
other_values = ["fe", "DC", "DO", "BX", "FO", "CP"]
header_table_one = ["Nome", "Categoria"]
months = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro",
          "Novembro", "Dezembro"]


def inject_vba_with_table_management(file_path):
    create_table2_code = '''
       Sub CreateTable2()
    Dim ws As Worksheet
    Dim mes As Integer
    Dim numDias As Integer
    Dim letrasLegenda As Variant
    Dim coluna As Integer
    Dim linha As Integer
    Dim total As Integer
    Dim dia As Integer
    Dim dataAtual As Date
    Dim ano As Integer
    Dim workbookName As String
    Dim yearPosition As Long
    Dim headerRange As Range
    Dim legendRange As Range
    Dim totalRange As Range
    Dim dataRange As Range

    ' Define as letras a serem contadas na legenda
    letrasLegenda = Array("M", "T", "N", "TG", "R", "m", "t")

    ' Determinar o ano a partir do nome do arquivo
    workbookName = ThisWorkbook.Name
    yearPosition = InStr(workbookName, "_") + 1
    ano = CInt(Mid(workbookName, yearPosition, 4))

    For Each ws In ThisWorkbook.Sheets
        mes = ws.Index ' Supondo que a folha de janeiro seja a primeira e assim por diante
        numDias = Day(DateSerial(ano, mes + 1, 0)) ' Número de dias no mês

        ' Cabeçalhos das duas primeiras colunas
        ws.Cells(4, 1).Value = "Legenda"
        ws.Cells(4, 2).Value = "Definição"

        ' Preencher a coluna da "Legenda" com as letras
        For linha = 5 To 5 + UBound(letrasLegenda)
            ws.Cells(linha, 1).Value = letrasLegenda(linha - 5)
        Next linha

        ' Cabeçalhos de cada dia do mês
        For dia = 1 To numDias
            dataAtual = DateSerial(ano, mes, dia)
            ws.Cells(4, 2 + dia).Value = dia & " (" & Format(dataAtual, "ddd") & ")"
        Next dia

        ' Contar ocorrências das letras e preencher a tabela
        For linha = 5 To 5 + UBound(letrasLegenda)
            For dia = 1 To numDias
                coluna = 3 + dia - 1
                total = 0
                For i = 12 To ws.Cells(ws.Rows.Count, coluna).End(xlUp).Row ' Contar da linha 12 para baixo
                    If ws.Cells(i, coluna).Value = ws.Cells(linha, 1).Value Then
                        total = total + 1
                    End If
                Next i
                ws.Cells(linha, coluna).Value = total
            Next dia
        Next linha

        ' FORMATAÇÃO:
        ' Cor de fundo para o cabeçalho
        Set headerRange = ws.Range(ws.Cells(4, 1), ws.Cells(4, 2 + numDias))
        headerRange.Interior.Color = RGB(237, 224, 209) ' Tom bege claro
        headerRange.Borders.LineStyle = xlContinuous
        headerRange.Borders.Weight = xlThin
        headerRange.Font.Bold = True
        headerRange.HorizontalAlignment = xlCenter

        ' Cor de fundo para a legenda e total
        Set legendRange = ws.Range(ws.Cells(5, 1), ws.Cells(5 + UBound(letrasLegenda), 1))
        legendRange.Interior.Color = RGB(237, 224, 209) ' Mesma cor bege claro para as legendas
        legendRange.Borders.LineStyle = xlContinuous
        legendRange.Borders.Weight = xlThin

        ' Formatar a coluna "Total"
        Set totalRange = ws.Range(ws.Cells(5, 2), ws.Cells(5 + UBound(letrasLegenda), 2))
        totalRange.Interior.Color = RGB(237, 224, 209)
        totalRange.Borders.LineStyle = xlContinuous
        totalRange.Borders.Weight = xlThin

        ' Formatar a parte de dados
        Set dataRange = ws.Range(ws.Cells(5, 3), ws.Cells(5 + UBound(letrasLegenda), 2 + numDias))
        dataRange.Borders.LineStyle = xlContinuous
        dataRange.Borders.Weight = xlThin
        dataRange.HorizontalAlignment = xlCenter

    Next ws
End Sub
       '''
    create_headers_code = '''
    Sub InserirCabecalho()
    ' msgbox "Começo do headercode"
    Dim ws As Worksheet
    Dim ano As String
    Dim mes As String
    Dim ultimaColuna As Integer
    Dim nomeLivro As String
    Dim posicaoUnderscore As Integer

    ' Obtém o nome do livro e extrai o ano (presume que o nome do livro está no formato "Livro_<ano>")
    nomeLivro = ThisWorkbook.Name
    posicaoUnderscore = InStrRev(nomeLivro, "_")
    ano = Mid(nomeLivro, posicaoUnderscore + 1, 4)

    ' Itera por todas as planilhas do livro
    For Each ws In ThisWorkbook.Sheets
        ' Define o mês com base no nome da planilha (presume-se que o nome da folha é o nome do mês)
        mes = UCase(ws.Name)

        ' Inserir espaço para a imagem na coluna A1
        ws.Cells(1, 1).Value = "" ' Deixe em branco para a imagem

        ' Inserir os textos no cabeçalho
        ws.Cells(1, 2).Value = "MAPA DE HORÁRIO DE TRABALHO"
        ws.Cells(2, 2).Value = "EQUIPA GERAL- V1 – PROVISÓRIO – PARA ANÁLISE"
        ws.Cells(3, 2).Value = "Horário Organizado em Turnos"

        ' Inserir "ANO" e "Mês" na parte direita
        ultimaColuna = 7 ' A última coluna que será usada para o ano e o mês (ajustar conforme necessário)
        ws.Cells(1, ultimaColuna).Value = "ANO"
        ws.Cells(1, ultimaColuna + 1).Value = ano
        ws.Cells(2, ultimaColuna).Value = "Mês"
        ws.Cells(2, ultimaColuna + 1).Value = mes

        ' Mesclar células para o cabeçalho e ajustar a largura
        ws.Range(ws.Cells(1, 2), ws.Cells(1, ultimaColuna - 1)).Merge
        ws.Range(ws.Cells(2, 2), ws.Cells(2, ultimaColuna - 1)).Merge
        ws.Range(ws.Cells(3, 2), ws.Cells(3, ultimaColuna - 1)).Merge

        ws.Cells(1, ultimaColuna).HorizontalAlignment = xlCenter
        ws.Cells(2, ultimaColuna).HorizontalAlignment = xlCenter

        ' Aplicar formatação de texto
        With ws.Cells(1, 2)
            .Font.Bold = True
            .HorizontalAlignment = xlCenter
        End With

        With ws.Cells(2, 2)
            .Font.Bold = True
            .HorizontalAlignment = xlCenter
        End With

        ws.Cells(3, 2).Font.Bold = True
        ws.Cells(3, 2).HorizontalAlignment = xlCenter

        ' Ajustar a largura das colunas conforme necessário
        ' ws.Columns(1).ColumnWidth = 20 ' Coluna para a imagem
        ' ws.Columns(2).ColumnWidth = 40 ' Coluna para o texto principal
        ' ws.Columns(ultimaColuna).ColumnWidth = 10 ' Coluna para "ANO"
        ' ws.Columns(ultimaColuna + 1).ColumnWidth = 15 ' Coluna para o ano e mês
    Next ws
End Sub
    '''
    create_table1_code = '''
    Sub CreateTable1()
        Dim ws As Worksheet
        Dim dynamicTableRange As Range
        Dim dynamicTable As ListObject
        Dim ano As Integer
        Dim mes As Integer
        Dim numDias As Integer
        Dim feriados As Variant
        Dim feriado As Variant
        Dim dia As Integer
        Dim dataAtual As Date
        Dim coluna As Integer
        Dim workbookName As String
        Dim yearPosition As Long
        Dim dynamicTableName As String
        Application.EnableEvents = True

        ' Determinar o ano a partir do nome do arquivo
        workbookName = ThisWorkbook.Name
        yearPosition = InStr(workbookName, "_") + 1
        ano = CInt(Mid(workbookName, yearPosition, 4))

        ' Definir os feriados
        feriados = Array(Array(1, 1), Array(25, 4), Array(1, 5), Array(10, 6), Array(15, 8), Array(5, 10), Array(1, 11), Array(1, 12), Array(8, 12), Array(25, 12))

        mes = 0
        For Each ws In ThisWorkbook.Sheets
            mes = mes + 1
            ' Determinar o mês a partir do nome da planilha
            numDias = Day(DateSerial(ano, mes + 1, 0))

            ' Definir o intervalo para a tabela dinâmica com as colunas apropriadas
            Set dynamicTableRange = ws.Range("A12").Resize(5, 2 + numDias + 3)

            ' Definir o nome único para a tabela dinâmica
            dynamicTableName = "DynamicTable_" & mes

            Set dynamicTable = ws.ListObjects.Add(SourceType:=xlSrcRange, Source:=dynamicTableRange, _
                XlListObjectHasHeaders:=xlYes, TableStyleName:="TableStyleMedium9")
            dynamicTable.Name = dynamicTableName
            dynamicTable.HeaderRowRange.Cells(1, 1).Value = "Nome"
            dynamicTable.HeaderRowRange.Cells(1, 2).Value = "Categoria"

            coluna = 3
            For dia = 1 To numDias
                dataAtual = DateSerial(ano, mes, dia)
                ws.Cells(12, coluna).Value = dia & " (" & Format(dataAtual, "ddd") & ")"
                If Weekday(dataAtual, vbMonday) >= 6 Then
                    ws.Cells(12, coluna).Interior.Color = RGB(255, 140, 0) ' Sábado
                ElseIf Weekday(dataAtual, vbMonday) = 7 Then
                    ws.Cells(12, coluna).Interior.Color = RGB(0, 0, 139) ' Domingo
                End If
                For Each feriado In feriados
                    If dia = feriado(0) And mes = feriado(1) Then
                        ws.Cells(12, coluna).Interior.Color = RGB(255, 0, 0) ' Feriado
                    End If
                Next feriado
                coluna = coluna + 1
            Next dia

            ' Adicionar cabeçalhos para "Horas Praticadas", "Horas Previstas" e "Diferença"
            ws.Cells(12, coluna).Value = "Horas Praticadas"
            ws.Cells(12, coluna + 1).Value = "Horas Previstas"
            ws.Cells(12, coluna + 2).Value = "Diferença"
        Next ws
    End Sub
    '''
    hours = '''
        Sub CalculateHours()
            Dim ws As Worksheet
            Dim userRow As Integer
            Dim lastCol As Integer
            Dim totalPraticadas As Double
            Dim i As Integer
            Dim horasPraticadasCol As Integer
            Dim hours As Integer
            Dim minutes As Integer
            Dim cellValue As Variant

            Dim validValues As Variant
            validValues = Array("M", "m", "T", "t", "N", "TG", "R")

            Dim redColor As Long
            redColor = RGB(255, 0, 0)

            For Each ws In ThisWorkbook.Worksheets
                lastCol = ws.Cells(12, ws.Columns.Count).End(xlToLeft).Column

                horasPraticadasCol = lastCol - 2

                For userRow = 13 To ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
                    If ws.Cells(userRow, 1).Value <> "" Then
                        totalPraticadas = 0

                        For i = 3 To lastCol - 3
                            If Not IsEmpty(ws.Cells(userRow, i).Value) Then
                                cellValue = ws.Cells(userRow, i).Value ' Mantém o valor original (case-sensitive)

                                Dim isValid As Boolean
                                isValid = False
                                Dim val As Variant
                                For Each val In validValues
                                    If cellValue = val Then
                                        isValid = True
                                        Exit For
                                    End If
                                Next val

                                Dim headerCellColor As Long
                                headerCellColor = ws.Cells(1, i).Interior.Color

                                If isValid And headerCellColor <> redColor Then
                                    totalPraticadas = totalPraticadas + 8
                                End If
                            End If
                        Next i

                        hours = Int(totalPraticadas)
                        minutes = (totalPraticadas - hours) * 60
                        ws.Cells(userRow, horasPraticadasCol).Value = hours & "h " & minutes & "m"
                    End If
                Next userRow
            Next ws
        End Sub

        Sub CalculateForecastedHours()
            Dim ws As Worksheet
            Dim userRow As Integer
            Dim lastCol As Integer
            Dim forecastedHours As Double
            Dim i As Integer
            Dim hoursPrevistasCol As Integer
            Dim totalDays As Integer
            Dim validCategories As Variant
            Dim categoryCell As String
            Dim isValidCategory As Boolean
            Dim hours As Integer
            Dim minutes As Integer

            validCategories = Array("AAD", "TA", "AC")

            Dim redColor As Long
            Dim orangeColor As Long
            Dim darkBlueColor As Long
            redColor = RGB(255, 0, 0)
            orangeColor = RGB(255, 165, 0)
            darkBlueColor = RGB(0, 0, 128)

            For Each ws In ThisWorkbook.Worksheets
                lastCol = ws.Cells(12, ws.Columns.Count).End(xlToLeft).Column

                hoursPrevistasCol = lastCol - 1

                For userRow = 13 To ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
                    If ws.Cells(userRow, 1).Value <> "" Then
                        categoryCell = ws.Cells(userRow, 2).Value
                        isValidCategory = False
                        For Each cat In validCategories
                            If categoryCell = cat Then
                                isValidCategory = True
                                Exit For
                            End If
                        Next cat

                        If isValidCategory Then
                            totalDays = 0
                            For i = 3 To lastCol - 3
                                Dim headerCellColor As Long
                                headerCellColor = ws.Cells(12, i).Interior.Color
                                If headerCellColor <> redColor And headerCellColor <> orangeColor Then
                                    totalDays = totalDays + 1
                                End If
                            Next i

                            If categoryCell = "AAD" Then
                                forecastedHours = 37 / 5 * totalDays
                            ElseIf categoryCell = "TA" Or categoryCell = "AC" Then
                                forecastedHours = 40 / 5 * totalDays
                            End If
                        Else
                            forecastedHours = 0
                        End If

                        hours = Int(forecastedHours)
                        minutes = (forecastedHours - hours) * 60
                        ws.Cells(userRow, hoursPrevistasCol).Value = hours & "h " & minutes & "m"
                    End If
                Next userRow
            Next ws
        End Sub

        Sub CalculateDifference()
            Dim ws As Worksheet
            Dim userRow As Integer
            Dim lastCol As Integer
            Dim diffCol As Integer
            Dim totalPraticadas As Double
            Dim forecastedHours As Double
            Dim i As Integer
            Dim horasPraticadasCol As Integer
            Dim hoursPrevistasCol As Integer
            Dim hours As Integer
            Dim minutes As Integer
            Dim difference As Double

            For Each ws In ThisWorkbook.Worksheets
                lastCol = ws.Cells(12, ws.Columns.Count).End(xlToLeft).Column

                horasPraticadasCol = lastCol - 2
                hoursPrevistasCol = lastCol - 1
                diffCol = lastCol

                For userRow = 13 To ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
                    If ws.Cells(userRow, 1).Value <> "" Then
                        Dim praticadas As String
                        Dim previstas As String
                        Dim praticadasHoras As Integer
                        Dim praticadasMinutos As Integer
                        Dim previstasHoras As Integer
                        Dim previstasMinutos As Integer
                        Dim praticadasTotalMin As Integer
                        Dim previstasTotalMin As Integer
                        Dim diffTotalMin As Integer

                        praticadas = Trim(ws.Cells(userRow, horasPraticadasCol).Value)
                        previstas = Trim(ws.Cells(userRow, hoursPrevistasCol).Value)

                        ' Inicialize variáveis para evitar erros caso a conversão falhe
                        praticadasHoras = 0
                        praticadasMinutos = 0
                        previstasHoras = 0
                        previstasMinutos = 0

                        ' Verificar se a string praticadas contém "h" e "m"
                        If InStr(praticadas, "h ") > 0 And InStr(praticadas, "m") > 0 Then
                            praticadasHoras = CInt(Split(praticadas, "h ")(0))
                            praticadasMinutos = CInt(Split(Split(praticadas, "h ")(1), "m")(0))
                        End If

                        ' Verificar se a string previstas contém "h" e "m"
                        If InStr(previstas, "h ") > 0 And InStr(previstas, "m") > 0 Then
                            previstasHoras = CInt(Split(previstas, "h ")(0))
                            previstasMinutos = CInt(Split(Split(previstas, "h ")(1), "m")(0))
                        End If

                        praticadasTotalMin = praticadasHoras * 60 + praticadasMinutos
                        previstasTotalMin = previstasHoras * 60 + previstasMinutos

                        diffTotalMin = praticadasTotalMin - previstasTotalMin


                         If praticadasTotalMin > previstasTotalMin Then
                                diffTotalMin = praticadasTotalMin - previstasTotalMin
                                hours = 1
                            Else
                                hours = -1
                                diffTotalMin = (praticadasTotalMin - previstasTotalMin) * (-1)
                            End If
                        hours = hours * Int(diffTotalMin / 60)
                        minutes = diffTotalMin Mod 60
                        ws.Cells(userRow, diffCol).Value = hours & "h " & minutes & "m"
                    End If
                Next userRow
            Next ws
        End Sub
        '''

    shift_counter =  '''
        Sub CountShifts()
    
        Dim ws As Worksheet
        Dim currCol As Long
        Dim currRow As Long
        Dim lastRow As Long
        Dim a As Long, b As Long, c As Long, d As Long, e As Long, f As Long, g As Long
        Dim value As Variant
        For Each ws In ThisWorkbook.Worksheets
            For currCol = 3 To ws.Cells(4, ws.Columns.Count).End(xlToLeft).Column
                ' Reset counters for each column
                a = 0
                b = 0
                c = 0
                d = 0
                e = 0
                f = 0
                g = 0
                lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
                For currRow = 13 To lastRow
                    value = ws.Cells(currRow, currCol).Value
                    Select Case value
                        Case "M"
                            a = a + 1
                            ' ws.Cells(currRow, currCol).Interior.Color = ws.Cells(currRow, currCol + 1).Interior.Color ' Match color of the cell to the right
                        Case "T"
                            b = b + 1
                            ' ws.Cells(currRow, currCol).Interior.Color = ws.Cells(currRow, currCol + 1).Interior.Color ' Match color of the cell to the right
                        Case "N"
                            c = c + 1
                            ' ws.Cells(currRow, currCol).Interior.Color = ws.Cells(currRow, currCol + 1).Interior.Color ' Match color of the cell to the right
                        Case "TG"
                            d = d + 1
                            ' ws.Cells(currRow, currCol).Interior.Color = ws.Cells(currRow, currCol - 1).Interior.Color ' Match color of the cell to the right
                        Case "R"
                            e = e + 1
                            ws.Cells(currRow, currCol).Interior.Color = ws.Cells(currRow, currCol - 1).Interior.Color ' Match color of the cell to the right
                        Case "m"
                            f = f + 1
                            ' ws.Cells(currRow, currCol).Interior.Color = ws.Cells(currRow, currCol - 1).Interior.Color ' Match color of the cell to the right
                        Case "t"
                            g = g + 1
                            ' ws.Cells(currRow, currCol).Interior.Color = ws.Cells(currRow, currCol - 1).Interior.Color ' Match color of the cell to the right
                        Case Is = "" ' Check for empty cell
                            ' ws.Cells(currRow, currCol).Interior.Color = ws.Cells(currRow, currCol - 1).Interior.Color ' Match color of the cell to the right
                        Case Else
                            MsgBox "Error: Invalid value '" & value & "' in cell " & ws.Cells(currRow, currCol).Address, vbExclamation
                            ' ws.Cells(currRow, currCol).Interior.Color = RGB(255, 0, 0) ' Paint the cell red
                    End Select
                Next currRow
                ws.Cells(5, currCol).Value = a
                ws.Cells(6, currCol).Value = b
                ws.Cells(7, currCol).Value = c
                ws.Cells(8, currCol).Value = d
                ws.Cells(9, currCol).Value = e
                ws.Cells(10, currCol).Value = f
                ws.Cells(11, currCol).Value = g
            Next currCol
        Next ws
    End Sub
    '''


    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        workbook = excel.Workbooks.Open(Filename=file_path)
        vba_module = workbook.VBProject.VBComponents.Add(1)
        vba_module.CodeModule.AddFromString(create_headers_code)
        vba_module.CodeModule.AddFromString(create_table1_code)
        vba_module.CodeModule.AddFromString(create_table2_code)
        vba_module.CodeModule.AddFromString(hours)
        vba_module.CodeModule.AddFromString(shift_counter)
        excel.application.run("CreateTable2")
        excel.application.run("InserirCabecalho")
        excel.application.run("CreateTable1")
        for sheet_name in months:
            ws = workbook.Sheets(sheet_name)
            code_module = workbook.VBProject.VBComponents(ws.CodeName).CodeModule
            worksheet_updater = f"""
               Private Sub Worksheet_Change(ByVal Target As Range)
                   On Error GoTo ExitHandler
                   Application.EnableEvents = False ' Disable events to prevent infinite loop

                   Dim lastLine As Long
                   Dim lastCol As Long
                   Dim monitoredRange As Range

                   lastLine = Cells(Rows.Count, 1).End(xlUp).Row ' Column A for last row
                   lastCol = Cells(12, Columns.Count).End(xlToLeft).Column - 3 ' Row 12 for last column minus 3

                   Set monitoredRange = Range(Cells(13, 1), Cells(lastLine, lastCol))

                   ' Check if the change is within the monitored range
                   If Not Intersect(Target, monitoredRange) Is Nothing Then
                       ' Call the three subs
                       Call CalculateHours
                       Call CalculateForecastedHours
                       Call CalculateDifference
                       Call CountShifts
                   End If

        ExitHandler:
                   Application.EnableEvents = True ' Re-enable events
               End Sub
               """
            code_module.AddFromString(worksheet_updater)
        workbook.Save()
        workbook.Close(SaveChanges=True)
        excel.Quit()

        print("VBA code injected and tables created successfully!")

    except Exception as e:
        print(f"Error injecting VBA code or creating tables: {e}")


def create_xlsm_file(directory, year):
    file_path = f"{directory}/Livro_{year}.xlsm"

    wb = Workbook()
    std_sheet = wb.active
    wb.remove(std_sheet)

    for month_num, month in enumerate(months, 1):
        wb.create_sheet(title=month)

    wb.save(file_path)

    # Open the workbook with VBA support enabled
    wb_vba = load_workbook(file_path, keep_vba=True)
    wb_vba.save(file_path)

    print(f"Workbook saved has {file_path}.")
    inject_vba_with_table_management(file_path)


def run():
    root = Tk()
    root.withdraw()

    year = simpledialog.askinteger("Year: ", "Insert a valid year number (ex: 2024):", minvalue=1900, maxvalue=2100)
    if year is None:
        return

    directory = filedialog.askdirectory(title="Choose the folder where to store your xlsm file")
    if directory:
        create_xlsm_file(directory, year)


run()